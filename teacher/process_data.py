import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import StringIO


def generate_excel(result_data, fig):
    # Create a Pandas DataFrame from the result_data
    data = [(result.student_id.name, result.student_id.prn, result.obtained_marks, result.TAG) for result in
            result_data]
    df = pd.DataFrame(data, columns=['Name', 'PRN', 'Marks', 'Type of Tag'])

    # Sort the DataFrame by the "Type of Tag" column
    df = df.sort_values(by='Type of Tag')

    # Create a new Excel workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Data'

    # Add the column names as the first row
    for col_num, column_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num, value=column_name)

    # Write the sorted DataFrame to the Excel worksheet
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Insert the Plotly figure into the Excel worksheet as an image
    fig.write_image('plot.png')  # Save the Plotly figure as an image (if not saved already)
    img = Image('plot.png')
    worksheet.add_image(img, 'E1')

    # Save the Excel file
    excel_file = 'output_excel_file.xlsx'
    workbook.save(excel_file)

    return excel_file


def generate_comparison_csv(result_data):
    # Create a string buffer to store the CSV data

    csv_buffer = StringIO()

    # Define the CSV fields
    field_names = ['Name', 'PRN', 'Marks', 'TAG']

    # Create a CSV writer using the string buffer
    csv_writer = csv.DictWriter(csv_buffer, fieldnames=field_names)

    # Write the header row
    csv_writer.writeheader()

    # Write the data rows
    for result in result_data:
        csv_writer.writerow({
            'Name': result.student_id.name,
            'PRN': result.student_id.prn,
            'Marks': result.obtained_marks,
            'TAG': result.TAG,
        })

    # Get the CSV data from the string buffer
    csv_data = csv_buffer.getvalue()

    # Close the string buffer
    csv_buffer.close()

    return csv_data
